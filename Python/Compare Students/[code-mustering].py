import openpyxl
import os

#Hàm lấy dssv đi học
def get_attendants_list(attendants_sheet):
    #Mở Workbook
    wb = openpyxl.load_workbook(attendants_sheet) 
    #Mở Sheet đang sử dụng tron Workbook
    sheet = wb.active 
    #Lấy giá trị max dòng, max cột
    max_row = sheet.max_row
    max_col = sheet.max_column
    #Tạo list attentdants
    attendants = []
    #Thêm các MSSV có đi học vào list
    for i in range(2,max_row):
        codeCell = sheet.cell(i, max_col-1)
        attendants.append(int(codeCell.value))
    #Trả về list
    return attendants


#Lấy đường dẫn trực tiếp của thư mục đang làm việc
file_path=os.path.dirname(__file__)
file_path=file_path.replace("\\",'/') + "/"

#Lấy đường dẫn trực tiếp của file attendant 
attendants_sheet = file_path+'attendants.xlsx'

#Lấy đường dẫn trực tiếp của file students
students_path = file_path + "DSSV.xlsx"

if __name__== "__main__":
    #Xóa màn hình để nhìn cho dễ
    os.system("cls")

    #Lấy danh sách những người đi học
    attendants = get_attendants_list(attendants_sheet)
    
    #Mở Workbook
    wb = openpyxl.load_workbook(students_path) 
    
    #Mở Sheet đang sử dụng tron Workbook
    sheet = wb.active 
    
    #Hỏi buổi để xác định cột điểm danh cần chỉnh sửa
    target_col = int(input('-Nhập buổi học: ')) + 4
    
    #In ra thông tin buổi học để kiểm tra
    print('------------------------------------------------')
    print(sheet.cell(row = 1, column = target_col).value)
    print('------------------------------------------------')
    
    #Hỏi có muốn điểm danh?
    choice = input("-Tiến Hành Điểm Danh? (y/n): " )
    if (choice == "n"):
        exit(1)
    print('------------------------------------------------')
    
    #Lấy giá trị max dòng
    max_row = sheet.max_row
    
    #Cập nhật giá trị điểm danh 
    for i in attendants:
        for j in range(2, max_row + 1): 
            #Lấy ô chứa MSSV
            codeCell = sheet.cell(row = j, column = 2)
            #Lấy ô chứa giá trị điểm danh
            statusCell = sheet.cell(row = j, column = target_col) 
            #Chỉ thực hiện điểm danh cho sv trong file student có MSSV trong list attedant
            if codeCell.value == i and statusCell.value != 'p' :
                statusCell.value = 'x'
                break
                    
        #Lưu lại cập nhật cho file
        wb.save(students_path)
    
    #Xóa cập nhật vừa rồi:
    choice = str(input('Bạn có muốn xóa thao tác điểm danh vừa rồi? (y/n): '))
    if choice == 'y':
        for j in range(2, max_row + 1): 
            #Lấy ô chứa giá trị điểm danh
            statusCell = sheet.cell(row = j, column = target_col) 
            if (statusCell.value != 'p'):
                statusCell.value = ' '
        #Lưu lại cập nhật cho file
        wb.save(students_path)
                