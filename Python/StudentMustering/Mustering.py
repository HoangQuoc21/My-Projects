import openpyxl
import os
from playsound import playsound

#Lấy đường dẫn trực tiếp của thư mục đang làm việc
file_path=os.path.dirname(__file__)
file_path=file_path.replace("\\",'/') + "/"

def speakName(i):
    name_path = file_path + 'Audio/'+ str(i) + '.mp3'
    playsound(name_path)

if __name__== "__main__":
    #Xóa màn hình để nhìn cho dễ
    os.system("cls")
    
   
    
    #Lấy đường dẫn trực tiếp của file sheet
    sheet_path = file_path + "DSSV.xlsx"
    
    #Mở Workbook
    wb = openpyxl.load_workbook(sheet_path) 
    
    #Mở Sheet đang sử dụng tron Workbook
    sheet = wb.active 
    
    #Hỏi buổi để xác định cột điểm danh cần chỉnh sửa
    target_col = int(input('-Nhập buổi học: ')) + 4
    
    #In ra thông tin buổi học để kiểm tra
    print('------------------------------------------------')
    print(sheet.cell(row = 1, column = target_col).value)
    print('------------------------------------------------')
    
    #Nhập stt sinh viên bắt đầu điểm danh
    start = int(input('-Nhập stt sinh viên bắt đầu điểm danh (1-148): '))
    print('------------------------------------------------')
    
    #Hỏi có muốn điểm danh?
    choice = input("-Tiến Hành Điểm Danh? (y/n): " )
    if (choice == "n"):
        exit(1)
    print('------------------------------------------------')
    
    #Lấy giá trị max dòng
    max_row = sheet.max_row - 5
    
    #Cập nhật giá trị điểm danh  
    #Duyệt qua các dòng
    for i in range(start + 1, max_row + 1): 
        #Lấy ô chứa MSSV
        codeCell = sheet.cell(row = i, column = 2)
        #Lấy ô chứa tên
        nameCell = sheet.cell(row = i, column = 3)
        #Lấy ô chứa giá trị điểm danh
        statusCell = sheet.cell(row = i, column = target_col) 
        
        #Chỉ thực hiện điểm danh cho sv có trạng thái điểm danh != 'p' (nếu trạng thái điểm danh là 'p' nghĩa là bạn này đã xin nghỉ có phép)
        if (statusCell.value != 'p'):
            #In ra thông tin sinh viên đang điểm danh
            print(i-1," ", codeCell.value, " ", nameCell.value,': ', end = ' ')
    
            choice = ''
            while choice != 'x' and choice != ' ':
                #Đọc Tên Sinh Viên
                speakName(i-1)
                
                choice = str(input(''))
                
                if choice == 'x':
                    #In thông báo ra console 
                    print('Có Mặt!')
                    #Phát âm tham có đi học
                    playsound(file_path + 'Audio/Co.mp3')
                    #Cập nhật giá trị điểm danh là x
                    statusCell.value = 'x'
                    break
                    
                elif choice == ' ':
                    #In thông báo ra console
                    print('Vắng')
                    
                    #Phát âm thanh vắng
                    playsound(file_path + 'Audio/Vang.mp3')
                    
                    #Cập nhật giá trị điểm danh là khoảng trống
                    statusCell.value = ' '
                    break
                
        #Lưu lại cập nhật cho file
        wb.save(sheet_path)
   