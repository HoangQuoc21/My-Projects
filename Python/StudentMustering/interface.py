import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os

#Lấy đường dẫn trực tiếp của thư mục đang làm việc
file_path=os.path.dirname(__file__)
file_path=file_path.replace("\\",'/') + "/"

#Lấy đường dẫn trực tiếp của file sheet
sheet_path = file_path + "students.xlsx"

def update_attendance(status):
    global current_index
    selected_student = students[current_index]
    attendance_list.insert(tk.END, f"{selected_student} - {status}")
    update_excel(selected_student, status)
    
    # Di chuyển tới sinh viên tiếp theo
    current_index += 1
    if current_index < len(students):
        show_student_info()
    else:
        label_students.config(text="Đã hoàn tất điểm danh")
        button_yes.config(state=tk.DISABLED)
        button_no.config(state=tk.DISABLED)
        messagebox.showinfo("Hoàn thành", "Đã điểm danh xong!")

def update_excel(student_name, status):
    try:
        workbook = load_workbook(sheet_path)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == student_name:
                row_index = row[0].row
                sheet.cell(row=row_index, column=2, value=status)
                break
        workbook.save(sheet_path)
    except Exception as e:
        print("Error updating data to Excel:", e)

def show_student_info():
    global current_index
    selected_student = students[current_index]
    label_students.config(text=f"Danh sách sinh viên ({current_index + 1}/{len(students)})")
    label_student_name.config(text=selected_student)

# Tạo cửa sổ đồ họa
root = tk.Tk()
root.title("Điểm danh sinh viên")



# Đọc danh sách sinh viên từ file Excel
try:
    workbook = load_workbook(sheet_path)
    sheet = workbook.active
    students = []
    #Lấy giá trị max dòng
    max_row = sheet.max_row
    for i in range(2, max_row + 1): 
        #Lấy ô chứa MSSV
        codeCell = sheet.cell(row = i, column = 2)
        #Lấy ô chứa tên
        nameCell = sheet.cell(row = i, column = 3)
        students.append(nameCell.value)
    current_index = 0
    show_student_info()
except Exception as e:
    print("Error loading data from Excel:", e)
    students = []
    current_index = -1

# Tạo các phần tử giao diện
label_students = tk.Label(root, text=f"Danh sách sinh viên ({current_index + 1}/{len(students)})")
label_students.pack()

label_student_name = tk.Label(root, text="")
label_student_name.pack()

button_yes = tk.Button(root, text="Có", command=lambda: update_attendance("Đi học"))
button_yes.pack()

button_no = tk.Button(root, text="Vắng", command=lambda: update_attendance("Vắng học"))
button_no.pack()

attendance_list = tk.Listbox(root)
attendance_list.pack()

if current_index == -1:
    button_yes.config(state=tk.DISABLED)
    button_no.config(state=tk.DISABLED)

# Chạy giao diện
root.mainloop()
